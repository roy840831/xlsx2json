import json
import argparse
from collections import OrderedDict
from openpyxl import Workbook,load_workbook
from pandas import DataFrame


def createXlsxFromJson(xlsxpath, jsobj):
    wb = Workbook()
    wb.remove(wb["Sheet"])
    global ws1
    ws1 = wb.create_sheet("json")

    analyzeObj(jsobj, row=1, column=1)

    wb.save(xlsxpath)

def readJson(jsonpath):
    with open(jsonpath, "r", encoding='utf-8') as file:
        jsobj = json.load(file)
    return jsobj

def analyzeObj(jsobj, row, column):

    #python dict equal json 
    if isinstance(jsobj, dict):
        row, column = analyzeDict(jsobj, row, column)
        return row, column

    elif isinstance(jsobj, list):

        if (
            not isinstance(jsobj[0], dict)
            or len(jsobj) < 2
            or not isinstance(jsobj[0], type(jsobj[1]))
        ):
            row, column = analyzeList(jsobj, row, column)
            return row, column

        listDictValue = next(iter(jsobj[0].items()))[1]
        if isinstance(listDictValue, dict) or isinstance(listDictValue, list):
            row, column = analyzeList(jsobj, row, column)
            return row, column

        sameKeyCount = [
            x for x in jsobj[0].keys() for y in jsobj[1].keys() if x == y
        ]
        if len(sameKeyCount) > 0:
            row, column = analyzeListDict(jsobj, row, column)
            return row, column

    else:
        ws1.cell(row=row, column=column).value = jsobj
        return row, column


def analyzeDict(jsobj, row, column):
    ws1.cell(row=row, column=column).value = "dict"
    for key in jsobj.keys():
        row += 1
        ws1.cell(row=row, column=column).value = key
        column += 1
        row, column = analyzeObj(jsobj[key], row, column)
        column -= 1

    return row, column


def analyzeList(jsobj, row, column):
    ws1.cell(row=row, column=column).value = "list"
    for index, value in enumerate(jsobj):
        row += 1
        ws1.cell(row=row, column=column).value = f"list{index}"
        column += 1
        row, column = analyzeObj(value, row, column)
        column -= 1

    return row, column


def analyzeListDict(jsobj, row, column):
    ws1.cell(row=row, column=column).value = "list-dict"
    df = DataFrame(jsobj)
    row += 1
    for index, columnsName in enumerate(df.columns):
        ws1.cell(row=row, column=column + index).value = columnsName
        for indexName in df[columnsName].index:
            ws1.cell(
                row=row + indexName + 1,
                column=column + index,
            ).value = df[columnsName][indexName]
    row += df.index[-1] + 1
    return row, column


def createJsonFromXlsx(jsonpath, xlsxpath):
    readXlsx(xlsxpath)

    global pyStructForJson
    pyStructForJson = {
        "dict": "object",
        "list": "array",
        "list-dict": "dictInArray",
    }

    match wsjson.cell(row=1, column=1).value:
        case "dict":
            jsonobj = checkDict(startRow=1, column=1, endRow=wsjson.max_row)
        case "list":
            jsonobj = list()
        case _:
            jsonobj = checkCellType

    # pprint(json.dumps(jsonobj))
    # Serializing json
    json_object = json.dumps(jsonobj, indent=4, ensure_ascii=False)
    with open(jsonpath, "w", encoding='utf8') as outfile:
        outfile.write(json_object)


def readXlsx(xlsxpath):
    wb = load_workbook(xlsxpath,data_only=True)
    global wsjson
    wsjson = wb["json"]


def checkCellType(row, lastRow, column):
    cellValue = wsjson.cell(row=row, column=column).value
    if cellValue is None:
        return None
    elif cellValue == "dict":
        return checkDict(startRow=row, column=column, endRow=lastRow)
    elif cellValue == "list":
        return checkList(startRow=row, column=column, endRow=lastRow)
    elif cellValue == "list-dict":
        return checkListDict(startRow=row, column=column, endRow=lastRow)
    else:
        return cellValue


def checkDict(startRow, column, endRow):
    # print(startRow,column,endRow)
    obj = OrderedDict()
    lastRow = endRow
    for row in range(endRow + 1, startRow, -1):
        cellValue = wsjson.cell(row=row, column=column).value
        if cellValue is not None and cellValue not in list(
            pyStructForJson.keys()
        ):
            obj[cellValue] = checkCellType(row, lastRow, column + 1)
            obj.move_to_end(cellValue, last=False)
            # print(row,lastRow,column,cellValue)
            lastRow = row - 1
    return obj


def checkList(startRow, column, endRow):
    # print(startRow,column,endRow)
    obj = list()
    lastRow = endRow
    for row in range(endRow + 1, startRow, -1):
        cellValue = wsjson.cell(row=row, column=column).value
        if cellValue is not None and cellValue not in list(
            pyStructForJson.keys()
        ):
            obj.insert(0, checkCellType(row, lastRow, column + 1))
            # print(row,lastRow,column,cellValue)
            lastRow = row - 1
    return obj

def checkListDict(startRow, column, endRow):
    # print(startRow,endRow,column)
    obj = list()
    for row in wsjson.iter_rows(
        min_row=startRow + 1,
        min_col=column,
        max_row=startRow + 1,
        values_only=True,
    ):
        columns = row

    for row in wsjson.iter_rows(
        min_row=startRow + 2,
        min_col=column,
        max_row=endRow,
        values_only=True,
    ):
        obj.append(row)

    df = DataFrame(obj, columns=columns)
    # remove empty row
    df = df.dropna(axis="index", how="all")
    # print(df)
    return df.to_dict("records")


if __name__ == "__main__":

    description = "description: Can create json file from xlsx "
    description += "or create xlsx file from json"

    paser = argparse.ArgumentParser(description=description)
    paser.add_argument("-x","--xlsx", type=str, help="xlsx file path(absolute or relative)")
    paser.add_argument("-j","--json", type=str, help="json file path(absolute or relative)")
    paser.add_argument("-o","--output", type=str, help="output file path(absolute or relative)")

    args = paser.parse_args()
    xlsxFilename = args.xlsx
    jsonFilename = args.json
    outputFile = args.output

    if jsonFilename is None and xlsxFilename is None:
        print("at least --json filepath or --xlxs filepath")
        print("use -h get help ")
    elif xlsxFilename is not None:
        if outputFile is None:
            outputFile = ".\\output.json"
        print(f"{xlsxFilename} to {outputFile}")
        createJsonFromXlsx(jsonpath=outputFile, xlsxpath=xlsxFilename)
    elif jsonFilename is not None:
        if outputFile is None:
            outputFile = ".\\output.xlsx"
        print(f"{jsonFilename} to {outputFile}")
        createXlsxFromJson(xlsxpath=outputFile, jsobj=readJson(jsonFilename))
    else:
        print("please just choose one kind file")
        print("use -h get help ")